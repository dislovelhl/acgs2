"""
XLSX document generator using openpyxl
Constitutional Hash: cdd01ef066bc6cf2
"""

import logging
from pathlib import Path
from typing import Any, Dict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from .base import BaseGenerator

logger = logging.getLogger(__name__)


class XLSXGenerator(BaseGenerator):
    """XLSX document generator for compliance documents"""

    def __init__(self, output_dir: str = "/app/documents"):
        super().__init__(output_dir)
        if not OPENPYXL_AVAILABLE:
            logger.warning("openpyxl not available, XLSX generation will fail")

    def generate(self, data: Dict[str, Any], filename: str) -> Path:
        """
        Generate an XLSX document from data.

        Args:
            data: Document data dictionary
            filename: Output filename (without extension)

        Returns:
            Path to generated XLSX file
        """
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl not available")

        output_path = self._get_output_path(filename, "xlsx")

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Compliance Data"

        # Add content based on document type
        doc_type = data.get("document_type", "generic")

        if doc_type == "risk_assessment":
            self._build_risk_assessment(ws, data)
        elif doc_type == "human_oversight":
            self._build_human_oversight(ws, data)
        elif doc_type == "compliance_checklist":
            self._build_compliance_checklist(ws, data)
        elif doc_type == "quarterly_report":
            self._build_quarterly_report(ws, data)
        else:
            self._build_generic(ws, data)

        # Save workbook
        wb.save(str(output_path))

        logger.info(f"Generated XLSX: {output_path}")
        return output_path

    def _build_risk_assessment(self, ws, data: Dict[str, Any]):
        """Build risk assessment spreadsheet."""
        row = 1

        # Header
        ws['A1'] = "Risk Assessment"
        ws['A1'].font = Font(bold=True, size=14)
        row += 2

        # Metadata
        ws[f'A{row}'] = "System:"
        ws[f'B{row}'] = data.get('system_name', 'N/A')
        row += 1
        ws[f'A{row}'] = "Assessment Date:"
        ws[f'B{row}'] = str(data.get('assessment_date', 'N/A'))
        row += 1
        ws[f'A{row}'] = "Assessor:"
        ws[f'B{row}'] = data.get('assessor_name', 'N/A')
        row += 2

        # Risk factors table
        risk_factors = data.get("risk_factors", [])
        if risk_factors:
            # Header row
            headers = ["Category", "Description", "Likelihood", "Impact", "Risk Level"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            row += 1

            # Data rows
            for factor in risk_factors:
                ws.cell(row=row, column=1).value = factor.get("category", "")
                ws.cell(row=row, column=2).value = factor.get("description", "")
                ws.cell(row=row, column=3).value = factor.get("likelihood", "")
                ws.cell(row=row, column=4).value = factor.get("impact", "")
                ws.cell(row=row, column=5).value = factor.get("risk_level", "")
                row += 1

            # Auto-adjust column widths
            for col in range(1, 6):
                ws.column_dimensions[get_column_letter(col)].width = 20

    def _build_human_oversight(self, ws, data: Dict[str, Any]):
        """Build human oversight spreadsheet."""
        row = 1

        # Header
        ws['A1'] = "Human Oversight Assessment"
        ws['A1'].font = Font(bold=True, size=14)
        row += 2

        # Metadata
        ws[f'A{row}'] = "System:"
        ws[f'B{row}'] = data.get('system_name', 'N/A')
        row += 2

        # Oversight measures table
        measures = data.get("oversight_measures", [])
        if measures:
            headers = ["Measure Type", "Description", "Responsible Role", "Frequency"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            row += 1

            for measure in measures:
                ws.cell(row=row, column=1).value = measure.get("measure_type", "")
                ws.cell(row=row, column=2).value = measure.get("description", "")
                ws.cell(row=row, column=3).value = measure.get("responsible_role", "")
                ws.cell(row=row, column=4).value = measure.get("frequency", "")
                row += 1

            for col in range(1, 5):
                ws.column_dimensions[get_column_letter(col)].width = 25

    def _build_compliance_checklist(self, ws, data: Dict[str, Any]):
        """Build compliance checklist spreadsheet."""
        row = 1

        # Header
        ws['A1'] = "Compliance Checklist"
        ws['A1'].font = Font(bold=True, size=14)
        row += 2

        # Metadata
        ws[f'A{row}'] = "System:"
        ws[f'B{row}'] = data.get('system_name', 'N/A')
        row += 1
        ws[f'A{row}'] = "Overall Status:"
        ws[f'B{row}'] = data.get('overall_status', 'N/A')
        row += 2

        # Findings table
        findings = data.get("findings", [])
        if findings:
            headers = ["Article", "Requirement", "Status", "Severity"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            row += 1

            for finding in findings:
                ws.cell(row=row, column=1).value = finding.get("article", "")
                ws.cell(row=row, column=2).value = finding.get("requirement", "")
                ws.cell(row=row, column=3).value = finding.get("status", "")
                ws.cell(row=row, column=4).value = finding.get("severity", "")
                row += 1

            for col in range(1, 5):
                ws.column_dimensions[get_column_letter(col)].width = 25

    def _build_quarterly_report(self, ws, data: Dict[str, Any]):
        """Build quarterly report spreadsheet."""
        row = 1

        # Header
        ws['A1'] = "Quarterly Compliance Report"
        ws['A1'].font = Font(bold=True, size=14)
        row += 2

        # Metrics
        metrics = data.get("report_period", {})
        if metrics:
            ws[f'A{row}'] = "Metric"
            ws[f'B{row}'] = "Value"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].font = Font(bold=True)
            row += 1

            metric_rows = [
                ("Total Assessments", metrics.get('total_assessments', 0)),
                ("Compliant Systems", metrics.get('compliant_systems', 0)),
                ("Non-Compliant Systems", metrics.get('non_compliant_systems', 0)),
                ("Critical Findings", metrics.get('critical_findings', 0)),
            ]

            for metric_name, metric_value in metric_rows:
                ws.cell(row=row, column=1).value = metric_name
                ws.cell(row=row, column=2).value = metric_value
                row += 1

            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 15

    def _build_generic(self, ws, data: Dict[str, Any]):
        """Build generic spreadsheet."""
        ws['A1'] = data.get("title", "Compliance Document")
        ws['A1'].font = Font(bold=True, size=14)

        content = data.get("content", "")
        if content:
            ws['A3'] = content
