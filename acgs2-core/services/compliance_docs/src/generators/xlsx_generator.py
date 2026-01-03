"""
XLSX Document Generator using openpyxl
"""

import logging
from typing import Any, Dict

from openpyxl import Workbook

logger = logging.getLogger(__name__)


class XLSXGenerator:
    """Generates XLSX compliance reports (evidence matrix)."""

    def generate(self, content: Dict[str, Any], output_path: str) -> str:
        """
        Generate XLSX report from content dictionary.

        Args:
            content: Dictionary containing report data
            output_path: Path to save the generated XLSX

        Returns:
            Path to the generated file
        """
        try:
            wb = Workbook(write_only=True)

            # Metadata Sheet
            ws_meta = wb.create_sheet("Metadata")
            ws_meta.append(["Field", "Value"])
            if "metadata" in content:
                for k, v in content["metadata"].items():
                    ws_meta.append([str(k), str(v)])

            # Evidence Sheet
            if "evidence" in content:
                ws_evidence = wb.create_sheet("Evidence")
                evidence_data = content["evidence"]
                if isinstance(evidence_data, list) and evidence_data:
                    # Headers (keys of first item)
                    headers = list(evidence_data[0].keys())
                    ws_evidence.append(headers)

                    for item in evidence_data:
                        row = [str(item.get(h, "")) for h in headers]
                        ws_evidence.append(row)

            # Additional Sections as Sheets
            if "sections" in content:
                for section in content["sections"]:
                    if "table" in section and section["table"]:
                        sheet_title = section.get("title", "Section")[:31]  # Excel limit
                        ws = wb.create_sheet(sheet_title)
                        for row in section["table"]:
                            ws.append([str(cell) for cell in row])

            wb.save(output_path)
            logger.info(f"XLSX generated successfully at {output_path}")
            return output_path

        except Exception as e:
            logger.error(f"Failed to generate XLSX: {e}")
            raise
