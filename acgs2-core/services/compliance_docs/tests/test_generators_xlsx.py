"""
Unit tests for XLSX document generator in the compliance-docs-service.

Tests cover:
- XLSX generator class initialization
- Custom styles for compliance spreadsheets
- Table builder functionality
- Evidence matrix generation for all four compliance frameworks (SOC 2, ISO 27001, GDPR, EU AI Act)
- Buffer generation for streaming responses
- Write-only mode for large files
- Error handling for unsupported frameworks
- Edge cases (minimal data, empty data)
- Output path handling
"""

import io
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from unittest.mock import patch

import pytest
from openpyxl import load_workbook

from generators.xlsx_generator import (
    ComplianceXLSXGenerator,
    ComplianceXLSXStyles,
    XLSXTableBuilder,
    generate_xlsx,
    generate_xlsx_to_buffer,
    _get_output_path,
    _ensure_output_dir,
    LARGE_FILE_THRESHOLD,
)
from models.base import ComplianceFramework


class TestComplianceXLSXStyles:
    """Tests for ComplianceXLSXStyles class."""

    def test_styles_initialization(self):
        """Test that styles are initialized correctly."""
        styles = ComplianceXLSXStyles()
        assert styles is not None

    def test_color_definitions(self):
        """Test that color definitions are present."""
        styles = ComplianceXLSXStyles()
        assert styles.PRIMARY_COLOR is not None
        assert styles.SECONDARY_COLOR is not None
        assert styles.HEADER_BG_COLOR is not None
        assert styles.SUCCESS_COLOR is not None
        assert styles.DANGER_COLOR is not None
        assert styles.WARNING_COLOR is not None

    def test_font_styles(self):
        """Test that font styles are created."""
        styles = ComplianceXLSXStyles()
        assert styles.title_font is not None
        assert styles.header_font is not None
        assert styles.body_font is not None
        assert styles.bold_font is not None
        assert styles.success_font is not None
        assert styles.danger_font is not None
        assert styles.warning_font is not None

    def test_fill_styles(self):
        """Test that fill styles are created."""
        styles = ComplianceXLSXStyles()
        assert styles.header_fill is not None
        assert styles.alt_row_fill is not None
        assert styles.success_fill is not None
        assert styles.danger_fill is not None
        assert styles.warning_fill is not None

    def test_border_style(self):
        """Test that border style is created."""
        styles = ComplianceXLSXStyles()
        assert styles.border is not None

    def test_alignment_styles(self):
        """Test that alignment styles are created."""
        styles = ComplianceXLSXStyles()
        assert styles.center_align is not None
        assert styles.left_align is not None
        assert styles.wrap_align is not None

    def test_get_status_style_compliant(self):
        """Test status style for compliant status."""
        styles = ComplianceXLSXStyles()
        font, fill = styles.get_status_style("compliant")
        assert font is styles.success_font
        assert fill is styles.success_fill

    def test_get_status_style_effective(self):
        """Test status style for effective status."""
        styles = ComplianceXLSXStyles()
        font, fill = styles.get_status_style("effective")
        assert font is styles.success_font
        assert fill is styles.success_fill

    def test_get_status_style_non_compliant(self):
        """Test status style for non-compliant status."""
        styles = ComplianceXLSXStyles()
        font, fill = styles.get_status_style("non-compliant")
        assert font is styles.danger_font
        assert fill is styles.danger_fill

    def test_get_status_style_failed(self):
        """Test status style for failed status."""
        styles = ComplianceXLSXStyles()
        font, fill = styles.get_status_style("failed")
        assert font is styles.danger_font
        assert fill is styles.danger_fill

    def test_get_status_style_pending(self):
        """Test status style for pending status."""
        styles = ComplianceXLSXStyles()
        font, fill = styles.get_status_style("pending")
        assert font is styles.warning_font
        assert fill is styles.warning_fill

    def test_get_status_style_in_progress(self):
        """Test status style for in_progress status."""
        styles = ComplianceXLSXStyles()
        font, fill = styles.get_status_style("in progress")
        assert font is styles.warning_font
        assert fill is styles.warning_fill

    def test_get_status_style_unknown(self):
        """Test status style for unknown status."""
        styles = ComplianceXLSXStyles()
        font, fill = styles.get_status_style("unknown")
        assert font is styles.body_font
        assert fill is None


class TestXLSXTableBuilder:
    """Tests for XLSXTableBuilder class."""

    def test_table_builder_initialization(self):
        """Test table builder initialization."""
        from openpyxl import Workbook

        wb = Workbook()
        styles = ComplianceXLSXStyles()
        builder = XLSXTableBuilder(wb, styles)
        assert builder is not None
        assert builder.workbook is wb
        assert builder.styles is styles
        assert builder.write_only is False

    def test_table_builder_write_only(self):
        """Test table builder in write-only mode."""
        from openpyxl import Workbook

        wb = Workbook(write_only=True)
        styles = ComplianceXLSXStyles()
        builder = XLSXTableBuilder(wb, styles, write_only=True)
        assert builder.write_only is True

    def test_create_evidence_sheet(self):
        """Test creating an evidence sheet."""
        from openpyxl import Workbook

        wb = Workbook()
        styles = ComplianceXLSXStyles()
        builder = XLSXTableBuilder(wb, styles)
        headers = ["ID", "Description", "Status"]
        rows = [["1", "Test", "Compliant"], ["2", "Test 2", "Pending"]]
        ws = builder.create_evidence_sheet("Test Sheet", headers, rows)
        assert ws is not None
        assert ws.title == "Test Sheet"
        # Check header row
        assert ws.cell(row=1, column=1).value == "ID"
        # Check data row
        assert ws.cell(row=2, column=1).value == "1"

    def test_create_evidence_sheet_with_status_column(self):
        """Test creating evidence sheet with status styling."""
        from openpyxl import Workbook

        wb = Workbook()
        styles = ComplianceXLSXStyles()
        builder = XLSXTableBuilder(wb, styles)
        headers = ["ID", "Status"]
        rows = [["1", "compliant"], ["2", "failed"]]
        ws = builder.create_evidence_sheet("Status Test", headers, rows, status_column=1)
        assert ws is not None

    def test_create_evidence_sheet_with_col_widths(self):
        """Test creating evidence sheet with column widths."""
        from openpyxl import Workbook

        wb = Workbook()
        styles = ComplianceXLSXStyles()
        builder = XLSXTableBuilder(wb, styles)
        headers = ["ID", "Name"]
        rows = [["1", "Test"]]
        ws = builder.create_evidence_sheet("Width Test", headers, rows, col_widths=[10, 30])
        assert ws is not None

    def test_create_evidence_sheet_freeze_panes(self):
        """Test that freeze panes are set on evidence sheet."""
        from openpyxl import Workbook

        wb = Workbook()
        styles = ComplianceXLSXStyles()
        builder = XLSXTableBuilder(wb, styles)
        headers = ["ID", "Name"]
        rows = [["1", "Test"]]
        ws = builder.create_evidence_sheet("Freeze Test", headers, rows)
        assert ws.freeze_panes == "A2"

    def test_create_summary_sheet(self):
        """Test creating a summary sheet."""
        from openpyxl import Workbook

        wb = Workbook()
        styles = ComplianceXLSXStyles()
        builder = XLSXTableBuilder(wb, styles)
        summary_data = [
            ("Report Type", "SOC 2"),
            ("Organization", "Test Org"),
            ("Date", "2024-06-15"),
        ]
        ws = builder.create_summary_sheet("Summary", summary_data)
        assert ws is not None
        assert ws.title == "Summary"
        assert ws.cell(row=1, column=1).value == "Report Type"
        assert ws.cell(row=1, column=2).value == "SOC 2"

    def test_create_evidence_sheet_write_only_mode(self):
        """Test creating evidence sheet in write-only mode."""
        from openpyxl import Workbook

        wb = Workbook(write_only=True)
        styles = ComplianceXLSXStyles()
        builder = XLSXTableBuilder(wb, styles, write_only=True)
        headers = ["ID", "Status"]
        rows = [["1", "OK"]]
        ws = builder.create_evidence_sheet("Write Only Test", headers, rows)
        assert ws is not None


class TestComplianceXLSXGenerator:
    """Tests for ComplianceXLSXGenerator class."""

    def test_generator_initialization(self):
        """Test XLSX generator initialization."""
        generator = ComplianceXLSXGenerator()
        assert generator is not None
        assert generator.write_only is False

    def test_generator_initialization_write_only(self):
        """Test XLSX generator with write-only mode."""
        generator = ComplianceXLSXGenerator(write_only=True)
        assert generator.write_only is True

    def test_create_workbook(self):
        """Test creating a new workbook."""
        generator = ComplianceXLSXGenerator()
        wb = generator._create_workbook()
        assert wb is not None
        assert generator._styles is not None
        assert generator._table_builder is not None

    def test_create_workbook_auto_write_only(self):
        """Test that large files trigger write-only mode."""
        generator = ComplianceXLSXGenerator()
        wb = generator._create_workbook(estimated_rows=LARGE_FILE_THRESHOLD + 1)
        assert wb is not None

    def test_format_date_with_datetime(self):
        """Test date formatting with datetime object."""
        generator = ComplianceXLSXGenerator()
        dt = datetime(2024, 6, 15, tzinfo=timezone.utc)
        result = generator._format_date(dt)
        # Excel dates are auto-converted, should return datetime
        assert result == dt

    def test_format_date_with_none(self):
        """Test date formatting with None value."""
        generator = ComplianceXLSXGenerator()
        result = generator._format_date(None)
        assert result == "N/A"

    def test_format_status(self):
        """Test status formatting."""
        generator = ComplianceXLSXGenerator()
        assert generator._format_status("in_progress") == "In Progress"
        assert generator._format_status("") == "N/A"
        assert generator._format_status(None) == "N/A"


class TestXLSXGeneratorSOC2:
    """Tests for SOC 2 XLSX matrix generation."""

    def test_generate_soc2_matrix(self, tmp_path, sample_soc2_xlsx_data):
        """Test generating a SOC 2 XLSX matrix."""
        output_path = tmp_path / "soc2_test.xlsx"
        generator = ComplianceXLSXGenerator()
        result = generator.generate_soc2_matrix(sample_soc2_xlsx_data, output_path)
        assert result == output_path
        assert output_path.exists()
        assert output_path.stat().st_size > 0

    def test_generate_soc2_matrix_opens_correctly(self, tmp_path, sample_soc2_xlsx_data):
        """Test that generated SOC 2 XLSX can be opened."""
        output_path = tmp_path / "soc2_test.xlsx"
        generator = ComplianceXLSXGenerator()
        generator.generate_soc2_matrix(sample_soc2_xlsx_data, output_path)
        wb = load_workbook(str(output_path))
        assert wb is not None
        assert "Summary" in wb.sheetnames

    def test_generate_soc2_matrix_has_summary(self, tmp_path, sample_soc2_xlsx_data):
        """Test that SOC 2 matrix has summary sheet."""
        output_path = tmp_path / "soc2_summary.xlsx"
        generator = ComplianceXLSXGenerator()
        generator.generate_soc2_matrix(sample_soc2_xlsx_data, output_path)
        wb = load_workbook(str(output_path))
        assert "Summary" in wb.sheetnames
        summary = wb["Summary"]
        assert summary.cell(row=1, column=1).value == "Report Type"

    def test_generate_soc2_matrix_has_evidence_sheet(self, tmp_path, sample_soc2_xlsx_data):
        """Test that SOC 2 matrix has evidence matrix sheet."""
        output_path = tmp_path / "soc2_evidence.xlsx"
        generator = ComplianceXLSXGenerator()
        generator.generate_soc2_matrix(sample_soc2_xlsx_data, output_path)
        wb = load_workbook(str(output_path))
        assert "Evidence Matrix" in wb.sheetnames

    def test_generate_soc2_matrix_with_mappings(self, tmp_path):
        """Test SOC 2 matrix with control mappings."""
        data = {
            "organization_name": "Test Org",
            "evidence_records": [],
            "control_mappings": [
                {
                    "soc2_control_id": "CC1.1",
                    "guardrail_control_id": "GR-001",
                    "guardrail_control_name": "Test Guardrail",
                    "mapping_rationale": "Test mapping",
                    "coverage_percentage": 100,
                    "gaps": [],
                },
            ],
        }
        output_path = tmp_path / "soc2_mappings.xlsx"
        generator = ComplianceXLSXGenerator()
        result = generator.generate_soc2_matrix(data, output_path)
        wb = load_workbook(str(result))
        assert "Control Mappings" in wb.sheetnames

    def test_generate_soc2_matrix_with_tsc_sections(self, tmp_path):
        """Test SOC 2 matrix with Trust Service Criteria sections."""
        data = {
            "organization_name": "Test Org",
            "evidence_records": [],
            "criteria_sections": [
                {
                    "criteria": "security",
                    "controls": [
                        {
                            "control_id": "CC1.1",
                            "title": "Test",
                            "control_objective": "Objective",
                            "implementation_guidance": "Guidance",
                        },
                    ],
                },
            ],
        }
        output_path = tmp_path / "soc2_tsc.xlsx"
        generator = ComplianceXLSXGenerator()
        result = generator.generate_soc2_matrix(data, output_path)
        wb = load_workbook(str(result))
        # Should have a TSC sheet (truncated to 31 chars)
        sheet_names = wb.sheetnames
        assert any("TSC" in name for name in sheet_names)

    def test_generate_soc2_matrix_default_path(self, sample_soc2_xlsx_data):
        """Test generating SOC 2 matrix with default output path."""
        generator = ComplianceXLSXGenerator()
        result = generator.generate_soc2_matrix(sample_soc2_xlsx_data)
        assert result.exists()
        assert "soc2_matrix" in result.name
        # Clean up
        result.unlink(missing_ok=True)


class TestXLSXGeneratorISO27001:
    """Tests for ISO 27001 XLSX matrix generation."""

    def test_generate_iso27001_matrix(self, tmp_path, sample_iso27001_xlsx_data):
        """Test generating an ISO 27001 XLSX matrix."""
        output_path = tmp_path / "iso27001_test.xlsx"
        generator = ComplianceXLSXGenerator()
        result = generator.generate_iso27001_matrix(sample_iso27001_xlsx_data, output_path)
        assert result == output_path
        assert output_path.exists()
        assert output_path.stat().st_size > 0

    def test_generate_iso27001_matrix_opens_correctly(self, tmp_path, sample_iso27001_xlsx_data):
        """Test that generated ISO 27001 XLSX can be opened."""
        output_path = tmp_path / "iso27001_test.xlsx"
        generator = ComplianceXLSXGenerator()
        generator.generate_iso27001_matrix(sample_iso27001_xlsx_data, output_path)
        wb = load_workbook(str(output_path))
        assert wb is not None
        assert "Summary" in wb.sheetnames

    def test_generate_iso27001_matrix_with_soa(self, tmp_path):
        """Test ISO 27001 matrix with Statement of Applicability."""
        data = {
            "organization_name": "Test Org",
            "isms_scope": "Test Scope",
            "statement_of_applicability": {
                "entries": [
                    {
                        "control_id": "A.5.1",
                        "control_title": "Security Policies",
                        "theme": "organizational",
                        "applicability": "applicable",
                        "justification": "Required",
                        "implementation_status": "implemented",
                        "implementation_method": "Policy",
                        "evidence_reference": "DOC-001",
                    },
                ],
            },
            "evidence_records": [],
        }
        output_path = tmp_path / "iso27001_soa.xlsx"
        generator = ComplianceXLSXGenerator()
        result = generator.generate_iso27001_matrix(data, output_path)
        wb = load_workbook(str(result))
        assert "Statement of Applicability" in wb.sheetnames

    def test_generate_iso27001_matrix_with_themes(self, tmp_path):
        """Test ISO 27001 matrix with theme sections."""
        data = {
            "organization_name": "Test Org",
            "evidence_records": [],
            "theme_sections": [
                {
                    "theme": "organizational",
                    "controls": [
                        {
                            "control_id": "A.5.1",
                            "title": "Policy",
                            "description": "Desc",
                            "status": "implemented",
                        },
                    ],
                },
            ],
        }
        output_path = tmp_path / "iso27001_themes.xlsx"
        generator = ComplianceXLSXGenerator()
        result = generator.generate_iso27001_matrix(data, output_path)
        wb = load_workbook(str(result))
        sheet_names = wb.sheetnames
        assert any("Organizational" in name for name in sheet_names)


class TestXLSXGeneratorGDPR:
    """Tests for GDPR XLSX matrix generation."""

    def test_generate_gdpr_matrix(self, tmp_path, sample_gdpr_xlsx_data):
        """Test generating a GDPR XLSX matrix."""
        output_path = tmp_path / "gdpr_test.xlsx"
        generator = ComplianceXLSXGenerator()
        result = generator.generate_gdpr_matrix(sample_gdpr_xlsx_data, output_path)
        assert result == output_path
        assert output_path.exists()
        assert output_path.stat().st_size > 0

    def test_generate_gdpr_matrix_opens_correctly(self, tmp_path, sample_gdpr_xlsx_data):
        """Test that generated GDPR XLSX can be opened."""
        output_path = tmp_path / "gdpr_test.xlsx"
        generator = ComplianceXLSXGenerator()
        generator.generate_gdpr_matrix(sample_gdpr_xlsx_data, output_path)
        wb = load_workbook(str(output_path))
        assert wb is not None
        assert "Summary" in wb.sheetnames

    def test_generate_gdpr_matrix_has_processing_activities(self, tmp_path, sample_gdpr_xlsx_data):
        """Test that GDPR matrix has processing activities sheet."""
        output_path = tmp_path / "gdpr_activities.xlsx"
        generator = ComplianceXLSXGenerator()
        generator.generate_gdpr_matrix(sample_gdpr_xlsx_data, output_path)
        wb = load_workbook(str(output_path))
        assert "Processing Activities" in wb.sheetnames

    def test_generate_gdpr_matrix_with_data_flows(self, tmp_path):
        """Test GDPR matrix with data flows."""
        data = {
            "organization_name": "Test Org",
            "controller_record": {"processing_activities": []},
            "data_flows": [
                {
                    "name": "Flow 1",
                    "flow_id": "DF-001",
                    "data_source": "Source",
                    "data_destination": "Dest",
                    "data_categories": ["Personal"],
                    "crosses_border": True,
                    "transfer_mechanism": "SCCs",
                    "encrypted_in_transit": True,
                },
            ],
        }
        output_path = tmp_path / "gdpr_flows.xlsx"
        generator = ComplianceXLSXGenerator()
        result = generator.generate_gdpr_matrix(data, output_path)
        wb = load_workbook(str(result))
        assert "Data Flows" in wb.sheetnames

    def test_generate_gdpr_matrix_with_security_measures(self, tmp_path):
        """Test GDPR matrix with security measures."""
        data = {
            "organization_name": "Test Org",
            "controller_record": {"processing_activities": []},
            "security_measures": [
                {
                    "measure_id": "SM-001",
                    "category": "technical",
                    "description": "Encryption at rest",
                    "status": "implemented",
                    "last_reviewed": datetime(2024, 6, 15, tzinfo=timezone.utc),
                },
            ],
        }
        output_path = tmp_path / "gdpr_security.xlsx"
        generator = ComplianceXLSXGenerator()
        result = generator.generate_gdpr_matrix(data, output_path)
        wb = load_workbook(str(result))
        assert "Security Measures" in wb.sheetnames


class TestXLSXGeneratorEUAIAct:
    """Tests for EU AI Act XLSX matrix generation."""

    def test_generate_euaiact_matrix(self, tmp_path, sample_euaiact_xlsx_data):
        """Test generating an EU AI Act XLSX matrix."""
        output_path = tmp_path / "euaiact_test.xlsx"
        generator = ComplianceXLSXGenerator()
        result = generator.generate_euaiact_matrix(sample_euaiact_xlsx_data, output_path)
        assert result == output_path
        assert output_path.exists()
        assert output_path.stat().st_size > 0

    def test_generate_euaiact_matrix_opens_correctly(self, tmp_path, sample_euaiact_xlsx_data):
        """Test that generated EU AI Act XLSX can be opened."""
        output_path = tmp_path / "euaiact_test.xlsx"
        generator = ComplianceXLSXGenerator()
        generator.generate_euaiact_matrix(sample_euaiact_xlsx_data, output_path)
        wb = load_workbook(str(output_path))
        assert wb is not None
        assert "Summary" in wb.sheetnames

    def test_generate_euaiact_matrix_has_ai_systems(self, tmp_path, sample_euaiact_xlsx_data):
        """Test that EU AI Act matrix has AI systems inventory sheet."""
        output_path = tmp_path / "euaiact_inventory.xlsx"
        generator = ComplianceXLSXGenerator()
        generator.generate_euaiact_matrix(sample_euaiact_xlsx_data, output_path)
        wb = load_workbook(str(output_path))
        assert "AI Systems Inventory" in wb.sheetnames

    def test_generate_euaiact_matrix_with_risk_assessments(self, tmp_path):
        """Test EU AI Act matrix with risk assessments."""
        data = {
            "organization_name": "Test Org",
            "ai_systems": [],
            "risk_assessments": [
                {
                    "assessment_id": "RA-001",
                    "system_name": "AI System",
                    "risk_level": "high_risk",
                    "high_risk_category": "Safety",
                    "assessment_date": datetime(2024, 6, 15, tzinfo=timezone.utc),
                    "assessor": "Test Assessor",
                    "key_findings": "Findings",
                    "recommendations": "Recommendations",
                },
            ],
        }
        output_path = tmp_path / "euaiact_risk.xlsx"
        generator = ComplianceXLSXGenerator()
        result = generator.generate_euaiact_matrix(data, output_path)
        wb = load_workbook(str(result))
        assert "Risk Assessments" in wb.sheetnames

    def test_generate_euaiact_matrix_with_conformity(self, tmp_path):
        """Test EU AI Act matrix with conformity assessments."""
        data = {
            "organization_name": "Test Org",
            "ai_systems": [],
            "conformity_assessments": [
                {
                    "system_id": "AIS-001",
                    "assessment_type": "internal_control",
                    "assessment_date": datetime(2024, 6, 15, tzinfo=timezone.utc),
                    "assessment_result": "passed",
                    "certificate_number": "CERT-001",
                    "expiry_date": datetime(2025, 6, 15, tzinfo=timezone.utc),
                    "notified_body": "Body",
                    "notes": "Notes",
                },
            ],
        }
        output_path = tmp_path / "euaiact_conformity.xlsx"
        generator = ComplianceXLSXGenerator()
        result = generator.generate_euaiact_matrix(data, output_path)
        wb = load_workbook(str(result))
        assert "Conformity Assessments" in wb.sheetnames

    def test_generate_euaiact_matrix_with_tech_docs(self, tmp_path):
        """Test EU AI Act matrix with technical documentation."""
        data = {
            "organization_name": "Test Org",
            "ai_systems": [],
            "technical_documentation": [
                {
                    "system_name": "AI System",
                    "documentation_status": "complete",
                    "annex_iv_requirements": {
                        "general_description": True,
                        "elements_development": True,
                        "monitoring_functioning": True,
                        "risk_management": False,
                        "data_governance": True,
                        "human_oversight": True,
                    },
                },
            ],
        }
        output_path = tmp_path / "euaiact_tech.xlsx"
        generator = ComplianceXLSXGenerator()
        result = generator.generate_euaiact_matrix(data, output_path)
        wb = load_workbook(str(result))
        assert "Technical Documentation" in wb.sheetnames


class TestGenerateXLSXFunction:
    """Tests for the generate_xlsx() main entry point function."""

    def test_generate_xlsx_soc2(self, tmp_path, sample_soc2_xlsx_data):
        """Test generate_xlsx with SOC 2 framework."""
        output_path = tmp_path / "soc2.xlsx"
        result = generate_xlsx(sample_soc2_xlsx_data, "soc2", output_path)
        assert result == output_path
        assert output_path.exists()

    def test_generate_xlsx_iso27001(self, tmp_path, sample_iso27001_xlsx_data):
        """Test generate_xlsx with ISO 27001 framework."""
        output_path = tmp_path / "iso27001.xlsx"
        result = generate_xlsx(sample_iso27001_xlsx_data, "iso27001", output_path)
        assert result == output_path
        assert output_path.exists()

    def test_generate_xlsx_gdpr(self, tmp_path, sample_gdpr_xlsx_data):
        """Test generate_xlsx with GDPR framework."""
        output_path = tmp_path / "gdpr.xlsx"
        result = generate_xlsx(sample_gdpr_xlsx_data, "gdpr", output_path)
        assert result == output_path
        assert output_path.exists()

    def test_generate_xlsx_euaiact(self, tmp_path, sample_euaiact_xlsx_data):
        """Test generate_xlsx with EU AI Act framework."""
        output_path = tmp_path / "euaiact.xlsx"
        result = generate_xlsx(sample_euaiact_xlsx_data, "euaiact", output_path)
        assert result == output_path
        assert output_path.exists()

    def test_generate_xlsx_with_framework_enum(self, tmp_path, sample_soc2_xlsx_data):
        """Test generate_xlsx with ComplianceFramework enum."""
        output_path = tmp_path / "soc2_enum.xlsx"
        result = generate_xlsx(sample_soc2_xlsx_data, ComplianceFramework.SOC2, output_path)
        assert result == output_path
        assert output_path.exists()

    def test_generate_xlsx_unsupported_framework(self, sample_soc2_xlsx_data):
        """Test generate_xlsx raises error for unsupported framework."""
        with pytest.raises(ValueError) as exc_info:
            generate_xlsx(sample_soc2_xlsx_data, "unsupported")
        assert "Unsupported framework" in str(exc_info.value)

    def test_generate_xlsx_case_insensitive_framework(self, tmp_path, sample_soc2_xlsx_data):
        """Test generate_xlsx is case insensitive for framework."""
        output_path = tmp_path / "soc2_upper.xlsx"
        result = generate_xlsx(sample_soc2_xlsx_data, "SOC2", output_path)
        assert result.exists()

    def test_generate_xlsx_write_only_mode(self, tmp_path, sample_soc2_xlsx_data):
        """Test generate_xlsx with write-only mode enabled."""
        output_path = tmp_path / "soc2_write_only.xlsx"
        result = generate_xlsx(sample_soc2_xlsx_data, "soc2", output_path, write_only=True)
        assert result.exists()


class TestGenerateXLSXToBuffer:
    """Tests for generate_xlsx_to_buffer() function."""

    def test_generate_xlsx_to_buffer_soc2(self, sample_soc2_xlsx_data):
        """Test generating SOC 2 XLSX to buffer."""
        buffer = generate_xlsx_to_buffer(sample_soc2_xlsx_data, "soc2")
        assert isinstance(buffer, io.BytesIO)
        content = buffer.getvalue()
        assert len(content) > 0
        # XLSX files are ZIP files, check for PK magic bytes
        assert content[:2] == b"PK"

    def test_generate_xlsx_to_buffer_iso27001(self, sample_iso27001_xlsx_data):
        """Test generating ISO 27001 XLSX to buffer."""
        buffer = generate_xlsx_to_buffer(sample_iso27001_xlsx_data, "iso27001")
        assert isinstance(buffer, io.BytesIO)
        assert buffer.getvalue()[:2] == b"PK"

    def test_generate_xlsx_to_buffer_gdpr(self, sample_gdpr_xlsx_data):
        """Test generating GDPR XLSX to buffer."""
        buffer = generate_xlsx_to_buffer(sample_gdpr_xlsx_data, "gdpr")
        assert isinstance(buffer, io.BytesIO)
        assert buffer.getvalue()[:2] == b"PK"

    def test_generate_xlsx_to_buffer_euaiact(self, sample_euaiact_xlsx_data):
        """Test generating EU AI Act XLSX to buffer."""
        buffer = generate_xlsx_to_buffer(sample_euaiact_xlsx_data, "euaiact")
        assert isinstance(buffer, io.BytesIO)
        assert buffer.getvalue()[:2] == b"PK"

    def test_generate_xlsx_to_buffer_unsupported_framework(self):
        """Test buffer generation raises error for unsupported framework."""
        with pytest.raises(ValueError) as exc_info:
            generate_xlsx_to_buffer({}, "unsupported")
        assert "Unsupported framework" in str(exc_info.value)

    def test_generate_xlsx_to_buffer_seek_position(self, sample_soc2_xlsx_data):
        """Test that buffer seek position is at start after generation."""
        buffer = generate_xlsx_to_buffer(sample_soc2_xlsx_data, "soc2")
        assert buffer.tell() == 0

    def test_generate_xlsx_to_buffer_can_be_opened(self, sample_soc2_xlsx_data):
        """Test that buffered XLSX can be opened."""
        buffer = generate_xlsx_to_buffer(sample_soc2_xlsx_data, "soc2")
        wb = load_workbook(buffer)
        assert wb is not None


class TestOutputPathHandling:
    """Tests for output path handling functions."""

    def test_get_output_path_default(self):
        """Test getting default output path."""
        path = _get_output_path()
        assert path is not None
        assert "compliance-reports" in str(path)

    def test_get_output_path_from_env(self):
        """Test getting output path from environment variable."""
        with patch.dict("os.environ", {"COMPLIANCE_OUTPUT_PATH": "/custom/path"}):
            path = _get_output_path()
            assert str(path) == "/custom/path"

    def test_ensure_output_dir(self, tmp_path):
        """Test ensuring output directory exists."""
        custom_path = tmp_path / "compliance-test"
        with patch.dict("os.environ", {"COMPLIANCE_OUTPUT_PATH": str(custom_path)}):
            result = _ensure_output_dir()
            assert result.exists()
            assert result.is_dir()


class TestEdgeCases:
    """Tests for edge cases and error handling."""

    def test_generate_xlsx_empty_data(self, tmp_path):
        """Test generating XLSX with empty data."""
        output_path = tmp_path / "empty.xlsx"
        result = generate_xlsx({}, "soc2", output_path)
        assert result.exists()

    def test_generate_xlsx_minimal_data(self, tmp_path, minimal_xlsx_data):
        """Test generating XLSX with minimal data."""
        output_path = tmp_path / "minimal.xlsx"
        result = generate_xlsx(minimal_xlsx_data, "soc2", output_path)
        assert result.exists()

    def test_generate_xlsx_none_values(self, tmp_path):
        """Test generating XLSX with None values in data."""
        data = {
            "organization_name": None,
            "evidence_records": [
                {"control_id": None, "description": None, "status": None},
            ],
        }
        output_path = tmp_path / "none_values.xlsx"
        result = generate_xlsx(data, "soc2", output_path)
        assert result.exists()

    def test_generate_xlsx_special_characters(self, tmp_path):
        """Test generating XLSX with special characters in data."""
        data = {
            "organization_name": "Test & Company <Corp>",
            "evidence_records": [
                {"control_id": "CC1.1", "description": 'Test & "Quote"'},
            ],
        }
        output_path = tmp_path / "special_chars.xlsx"
        result = generate_xlsx(data, "soc2", output_path)
        assert result.exists()

    def test_generate_xlsx_unicode_characters(self, tmp_path):
        """Test generating XLSX with unicode characters."""
        data = {
            "organization_name": "Test Org - Societe Francaise",
            "evidence_records": [],
        }
        output_path = tmp_path / "unicode.xlsx"
        result = generate_xlsx(data, "soc2", output_path)
        assert result.exists()

    def test_generate_xlsx_long_sheet_name(self, tmp_path):
        """Test that long sheet names are truncated to 31 characters."""
        data = {
            "organization_name": "Test Org",
            "evidence_records": [],
            "criteria_sections": [
                {
                    "criteria": "this_is_a_very_long_criteria_name_that_exceeds_31_chars",
                    "controls": [{"control_id": "1", "title": "T", "control_objective": "O"}],
                },
            ],
        }
        output_path = tmp_path / "long_name.xlsx"
        result = generate_xlsx(data, "soc2", output_path)
        wb = load_workbook(str(result))
        for sheet_name in wb.sheetnames:
            assert len(sheet_name) <= 31


class TestLargeFileThreshold:
    """Tests for large file threshold and write-only mode."""

    def test_large_file_threshold_value(self):
        """Test that LARGE_FILE_THRESHOLD is set correctly."""
        assert LARGE_FILE_THRESHOLD == 10000

    def test_auto_write_only_for_large_files(self):
        """Test that write-only mode is auto-enabled for large files."""
        generator = ComplianceXLSXGenerator()
        # This should trigger write-only mode
        wb = generator._create_workbook(estimated_rows=LARGE_FILE_THRESHOLD + 100)
        # The workbook should be created (we can't easily check write_only property)
        assert wb is not None


# Fixtures for XLSX generator tests
@pytest.fixture
def sample_soc2_xlsx_data():
    """Sample SOC 2 data for XLSX testing."""
    return {
        "organization_name": "ACGS Test Corporation",
        "audit_period_start": datetime(2024, 1, 1, tzinfo=timezone.utc),
        "audit_period_end": datetime(2024, 12, 31, tzinfo=timezone.utc),
        "total_controls": 50,
        "controls_tested": 45,
        "controls_effective": 42,
        "controls_with_exceptions": 3,
        "evidence_records": [
            {
                "control_id": "CC1.1",
                "criteria": "security",
                "title": "Commitment to Integrity",
                "description": "Test evidence",
                "evidence_type": "document",
                "collected_at": datetime(2024, 6, 15, tzinfo=timezone.utc),
                "design_effectiveness": "effective",
                "operating_effectiveness": "effective",
                "exceptions_noted": 0,
                "status": "compliant",
            },
        ],
    }


@pytest.fixture
def sample_iso27001_xlsx_data():
    """Sample ISO 27001 data for XLSX testing."""
    return {
        "organization_name": "ACGS Test Corporation",
        "isms_scope": "AI Guardrails Platform",
        "total_controls": 93,
        "applicable_controls": 80,
        "implemented_controls": 75,
        "implementation_percentage": 93.75,
        "statement_of_applicability": {
            "entries": [
                {
                    "control_id": "A.5.1",
                    "control_title": "Information Security Policy",
                    "theme": "organizational",
                    "applicability": "applicable",
                    "implementation_status": "implemented",
                },
            ],
        },
        "evidence_records": [
            {
                "control_id": "A.5.1",
                "theme": "organizational",
                "description": "Security policy evidence",
                "evidence_type": "document",
                "source": "Policy repository",
                "collected_at": datetime(2024, 6, 15, tzinfo=timezone.utc),
                "status": "compliant",
                "notes": "Reviewed Q2 2024",
            },
        ],
    }


@pytest.fixture
def sample_gdpr_xlsx_data():
    """Sample GDPR data for XLSX testing."""
    return {
        "organization_name": "ACGS Test Corporation",
        "entity_role": "controller",
        "controller_record": {
            "controller_name": "ACGS Test Corporation",
            "dpo": {"name": "Jane Doe", "email": "dpo@test.com"},
            "processing_activities": [
                {
                    "name": "AI Model Training",
                    "processing_id": "PA-001",
                    "purposes": ["Training", "Evaluation"],
                    "lawful_basis": "legitimate_interest",
                    "data_subject_categories": ["Employees"],
                    "personal_data_categories": ["Professional"],
                    "recipients": [{"name": "Cloud Provider"}],
                    "third_country_transfers": False,
                    "retention_period": "2 years",
                    "security_measures_summary": "Encryption",
                    "status": "active",
                },
            ],
        },
    }


@pytest.fixture
def sample_euaiact_xlsx_data():
    """Sample EU AI Act data for XLSX testing."""
    return {
        "organization_name": "ACGS Test Corporation",
        "organization_role": "provider",
        "high_risk_systems_count": 1,
        "limited_risk_systems_count": 2,
        "minimal_risk_systems_count": 0,
        "ai_systems": [
            {
                "system_id": "AIS-001",
                "system_name": "Guardrail Engine",
                "risk_level": "high_risk",
                "high_risk_category": "Safety Component",
                "intended_purpose": "Automated compliance checking",
                "provider_role": "provider",
                "compliance_status": "compliant",
                "last_assessment_date": datetime(2024, 6, 15, tzinfo=timezone.utc),
            },
        ],
    }


@pytest.fixture
def minimal_xlsx_data():
    """Minimal data for XLSX testing."""
    return {"organization_name": "Test Org"}
