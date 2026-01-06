"""
XLSX Styles for Compliance Document Generation.
Constitutional Hash: cdd01ef066bc6cf2
"""

from typing import Optional

from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)


class ComplianceXLSXStyles:
    """
    Custom styles for compliance XLSX documents.
    Provides professional, consistent styling across all compliance spreadsheets.
    """

    PRIMARY_COLOR = "1A365D"
    SECONDARY_COLOR = "2C5282"
    HEADER_BG_COLOR = "2C5282"
    HEADER_TEXT_COLOR = "FFFFFF"
    ALT_ROW_COLOR = "F7FAFC"
    SUCCESS_COLOR = "276749"
    DANGER_COLOR = "C53030"
    WARNING_COLOR = "C05621"
    BORDER_COLOR = "E2E8F0"

    def __init__(self) -> None:
        self._create_styles()

    def _create_styles(self) -> None:
        self.title_font = Font(name="Calibri", size=16, bold=True, color=self.PRIMARY_COLOR)
        self.header_font = Font(name="Calibri", size=11, bold=True, color=self.HEADER_TEXT_COLOR)
        self.subheader_font = Font(name="Calibri", size=12, bold=True, color=self.SECONDARY_COLOR)
        self.body_font = Font(name="Calibri", size=10, color="2D3748")
        self.bold_font = Font(name="Calibri", size=10, bold=True, color="2D3748")
        self.success_font = Font(name="Calibri", size=10, bold=True, color=self.SUCCESS_COLOR)
        self.danger_font = Font(name="Calibri", size=10, bold=True, color=self.DANGER_COLOR)
        self.warning_font = Font(name="Calibri", size=10, bold=True, color=self.WARNING_COLOR)

        self.header_fill = PatternFill(
            start_color=self.HEADER_BG_COLOR, end_color=self.HEADER_BG_COLOR, fill_type="solid"
        )
        self.alt_row_fill = PatternFill(
            start_color=self.ALT_ROW_COLOR, end_color=self.ALT_ROW_COLOR, fill_type="solid"
        )
        self.success_fill = PatternFill(start_color="C6F6D5", end_color="C6F6D5", fill_type="solid")
        self.danger_fill = PatternFill(start_color="FED7D7", end_color="FED7D7", fill_type="solid")
        self.warning_fill = PatternFill(start_color="FEEBC8", end_color="FEEBC8", fill_type="solid")

        thin_border = Side(style="thin", color=self.BORDER_COLOR)
        self.border = Border(
            left=thin_border, right=thin_border, top=thin_border, bottom=thin_border
        )

        self.center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        self.wrap_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

    def get_status_style(self, status: str) -> tuple[Font, Optional[PatternFill]]:
        status_lower = str(status).lower().replace("_", " ")
        words = status_lower.split()

        if (
            any(word in ["ineffective", "non-compliant", "failed", "no"] for word in words)
            or "non-compliant" in status_lower
        ):
            return self.danger_font, self.danger_fill
        elif any(
            word in ["effective", "compliant", "completed", "passed", "yes"] for word in words
        ):
            return self.success_font, self.success_fill
        elif (
            any(word in ["partial", "in-progress", "pending", "review"] for word in words)
            or "in progress" in status_lower
        ):
            return self.warning_font, self.warning_fill
        return self.body_font, None
