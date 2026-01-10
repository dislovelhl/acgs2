"""Constitutional Hash: cdd01ef066bc6cf2
XLSX Document Generator for Compliance Documentation Service

Generates professional compliance evidence matrices and reports in XLSX format
using openpyxl. Supports all compliance frameworks: SOC 2, ISO 27001, GDPR,
and EU AI Act.

This module uses write_only mode for large files (>10k rows) to optimize
memory usage. Python datetime objects are automatically converted to Excel
dates.

Note: Excel uses 1-indexed rows/columns in user-facing documentation.
"""

import io
import logging
import os
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, BinaryIO, Dict, List, Optional, Union

try:
    from src.core.shared.types import DocumentData, JSONDict, JSONValue
except ImportError:
    # Fallback for when src.core.shared.types is not available (e.g., in some test environments)
    # These definitions should ideally come from a central types file.
    JSONPrimitive = Union[str, int, float, bool, None]
    JSONDict = Dict[str, Any]
    JSONList = List[Any]
    JSONValue = Union[JSONPrimitive, JSONDict, JSONList]
    DocumentData = Dict[str, JSONValue]

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..models.base import ComplianceFramework
from .base import BaseGenerator

logger = logging.getLogger(__name__)

# Default output path for generated XLSX files
_DEFAULT_OUTPUT_PATH = Path(tempfile.gettempdir()) / "compliance-reports"

# Threshold for using write_only mode (10,000 rows)
LARGE_FILE_THRESHOLD = 10000


def _get_output_path() -> Path:
    """
    Resolve output path from environment or use default.

    Returns:
        Path to output directory for generated XLSX files.
    """
    output_path = os.getenv("COMPLIANCE_OUTPUT_PATH")
    if output_path:
        return Path(output_path)
    return _DEFAULT_OUTPUT_PATH


def _ensure_output_dir() -> Path:
    """
    Ensure the output directory exists.

    Returns:
        Path to the output directory.
    """
    output_path = _get_output_path()
    output_path.mkdir(parents=True, exist_ok=True)
    return output_path


# ruff: noqa: E402
from .xlsx_styles import ComplianceXLSXStyles


class XLSXTableBuilder:
    """
    Builder class for creating formatted tables in XLSX worksheets.

    Provides methods for creating compliance-specific tables with
    consistent styling and formatting.
    """

    def __init__(
        self,
        workbook: Workbook,
        styles: ComplianceXLSXStyles,
        write_only: bool = False,
    ) -> None:
        """
        Initialize the table builder.

        Args:
            workbook: The Workbook instance to add sheets to.
            styles: The XLSX styles instance to use.
            write_only: Whether workbook is in write_only mode.
        """
        self.workbook = workbook
        self.styles = styles
        self.write_only = write_only

    def create_evidence_sheet(
        self,
        sheet_name: str,
        headers: List[str],
        rows: List[List[JSONValue]],
        col_widths: Optional[List[int]] = None,
        status_column: Optional[int] = None,
    ) -> Worksheet:
        """
        Create a worksheet with evidence data.

        Args:
            sheet_name: Name for the worksheet.
            headers: List of column headers.
            rows: List of rows, each row is a list of cell values.
            col_widths: Optional list of column widths.
            status_column: Optional column index (0-based) for status styling.

        Returns:
            The created Worksheet.
        """
        if self.write_only:
            ws = self.workbook.create_sheet(sheet_name)
            # In write_only mode, append rows directly
            ws.append(headers)
            for row in rows:
                ws.append(row)
            return ws

        # Normal mode with full styling
        ws = self.workbook.create_sheet(sheet_name)

        # Add header row
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.styles.header_font
            cell.fill = self.styles.header_fill
            cell.alignment = self.styles.center_align
            cell.border = self.styles.border

        # Add data rows
        for row_idx, row_data in enumerate(rows, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = self.styles.body_font
                cell.alignment = self.styles.left_align
                cell.border = self.styles.border

                # Apply alternating row colors
                if row_idx % 2 == 0:
                    cell.fill = self.styles.alt_row_fill

                # Apply status styling
                if status_column is not None and col_idx == status_column + 1:
                    status_font, status_fill = self.styles.get_status_style(
                        str(value) if value else ""
                    )
                    cell.font = status_font
                    if status_fill:
                        cell.fill = status_fill

        # Set column widths
        if col_widths:
            for col_idx, width in enumerate(col_widths, start=1):
                ws.column_dimensions[get_column_letter(col_idx)].width = width
        else:
            # Auto-size based on headers
            for col_idx, header in enumerate(headers, start=1):
                ws.column_dimensions[get_column_letter(col_idx)].width = max(
                    len(str(header)) + 4, 12
                )

        # Freeze header row
        ws.freeze_panes = "A2"

        return ws

    def create_summary_sheet(
        self,
        sheet_name: str,
        summary_data: List[tuple[str, JSONValue]],
    ) -> Worksheet:
        """
        Create a summary sheet with key-value pairs.

        Args:
            sheet_name: Name for the worksheet.
            summary_data: List of (label, value) tuples.

        Returns:
            The created Worksheet.
        """
        if self.write_only:
            ws = self.workbook.create_sheet(sheet_name)
            for label, value in summary_data:
                ws.append([label, value])
            return ws

        ws = self.workbook.create_sheet(sheet_name)

        for row_idx, (label, value) in enumerate(summary_data, start=1):
            # Label cell
            label_cell = ws.cell(row=row_idx, column=1, value=label)
            label_cell.font = self.styles.bold_font
            label_cell.alignment = self.styles.left_align
            label_cell.border = self.styles.border

            # Value cell
            value_cell = ws.cell(row=row_idx, column=2, value=value)
            value_cell.font = self.styles.body_font
            value_cell.alignment = self.styles.left_align
            value_cell.border = self.styles.border

        # Set column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 50

        return ws


class XLSXGenerator(BaseGenerator):
    """
    Main XLSX generator for compliance documentation.

    Generates professional compliance evidence matrices and reports with
    consistent styling, proper structure, and support for all frameworks.
    """

    def __init__(
        self,
        output_dir: Union[str, Path, None] = None,
        write_only: bool = False,
    ) -> None:
        """
        Initialize the XLSX generator.

        Args:
            output_dir: Directory for generated documents.
            write_only: Whether to use write-only mode for large files.
        """
        if output_dir is None:
            output_dir = _get_output_path()
        super().__init__(str(output_dir))
        self.write_only = write_only
        self._workbook: Optional[Workbook] = None
        self._styles: Optional[ComplianceXLSXStyles] = None
        self._table_builder: Optional[XLSXTableBuilder] = None

    def generate(
        self,
        data: DocumentData,
        output_path: Optional[Union[str, Path]] = None,
    ) -> Path:
        """
        Produce a compliance matrix using a unified interface.

        Args:
            data: The document data, including a mandatory "document_type" key.
            output_path: Where to save the resulting XLSX.

        Returns:
            The Path where the file was written.

        Raises:
            ValueError: If document_type is missing or unsupported.
        """
        doc_type = data.get("document_type")
        if not doc_type:
            raise ValueError("Direct generation requires 'document_type' in data.")

        # Handle output path resolution if provided
        if output_path:
            output_path = Path(output_path)
            # Ensure extension
            if not output_path.suffix:
                output_path = output_path.with_suffix(".xlsx")
            # Ensure generated in output_dir if not absolute
            if not output_path.is_absolute():
                output_path = self.output_dir / output_path

        if doc_type == "soc2_report" or doc_type == "soc2_matrix":
            return self.generate_soc2_matrix(data, output_path)
        elif doc_type == "iso27001_report" or doc_type == "iso27001_matrix":
            return self.generate_iso27001_matrix(data, output_path)
        elif doc_type == "gdpr_report" or doc_type == "gdpr_matrix":
            return self.generate_gdpr_matrix(data, output_path)
        elif doc_type == "euaiact_report" or doc_type == "euaiact_matrix":
            return self.generate_euaiact_matrix(data, output_path)
        # Handle EU AI Act specific types from simpler generators
        elif doc_type in [
            "risk_assessment",
            "human_oversight",
            "compliance_checklist",
            "quarterly_report",
        ]:
            return self.generate_euaiact_matrix(data, output_path)
        else:
            return self.generate_euaiact_matrix(data, output_path)

    @property
    def workbook(self) -> Workbook:
        """Get the active Workbook instance."""
        from typing import cast

        return cast(Workbook, self._workbook)

    @property
    def table_builder(self) -> XLSXTableBuilder:
        """Get the active XLSXTableBuilder instance."""
        from typing import cast

        return cast(XLSXTableBuilder, self._table_builder)

    def _create_workbook(self, estimated_rows: int = 0) -> Workbook:
        """
        Create a new Workbook instance.

        Args:
            estimated_rows: Estimated number of rows for auto write_only decision.

        Returns:
            A new Workbook.
        """
        # Auto-enable write_only for large files
        use_write_only = self.write_only or estimated_rows > LARGE_FILE_THRESHOLD

        self._workbook = Workbook(write_only=use_write_only)
        self._styles = ComplianceXLSXStyles()
        self._table_builder = XLSXTableBuilder(
            self._workbook, self._styles, write_only=use_write_only
        )

        # Remove default sheet in normal mode
        if not use_write_only and "Sheet" in self._workbook.sheetnames:
            del self._workbook["Sheet"]

        return self._workbook

    @staticmethod
    def _format_date(value: Optional[Union[datetime, str]]) -> Union[datetime, str, None]:
        """
        Format a datetime value for Excel.

        Note: openpyxl auto-converts Python datetime to Excel dates.

        Args:
            value: Value to format.

        Returns:
            Formatted value (datetime or string).
        """
        if value is None:
            return "N/A"
        if isinstance(value, datetime):
            return value.replace(tzinfo=None)  # Excel doesn't support timezones
        return str(value)

    def _format_status(self, status: JSONValue) -> str:
        """Format a status value for display."""
        if not status:
            return "N/A"
        return str(status).replace("_", " ").title()

    def _save_workbook(
        self,
        output: Union[str, Path, BinaryIO],
    ) -> None:
        """
        Save the workbook to a file or buffer.

        Args:
            output: Output file path or file-like object.
        """
        if isinstance(output, (str, Path)):
            output_path = Path(output)
            output_path.parent.mkdir(parents=True, exist_ok=True)
            self.workbook.save(str(output_path))
        else:
            self.workbook.save(output)

    def generate_soc2_matrix(
        self,
        report_data: DocumentData,
        output_path: Optional[Union[str, Path]] = None,
    ) -> Path:
        """
        Generate a SOC 2 evidence matrix in XLSX format.

        Args:
            report_data: SOC 2 report data dictionary.
            output_path: Optional output file path.

        Returns:
            Path to the generated XLSX file.
        """
        evidence_records = report_data.get("evidence_records", [])
        self._create_workbook(len(evidence_records))

        org_name = report_data.get("organization_name", "Organization")
        audit_start = report_data.get("audit_period_start")
        audit_end = report_data.get("audit_period_end")

        # Summary sheet
        summary_data = [
            ("Report Type", "SOC 2 Type II Evidence Matrix"),
            ("Organization", org_name),
            ("Audit Period Start", self._format_date(audit_start)),
            ("Audit Period End", self._format_date(audit_end)),
            ("Generated At", self._format_date(datetime.now(timezone.utc))),
            ("Total Controls", report_data.get("total_controls", 0)),
            ("Controls Tested", report_data.get("controls_tested", 0)),
            ("Controls Effective", report_data.get("controls_effective", 0)),
            ("Controls with Exceptions", report_data.get("controls_with_exceptions", 0)),
        ]
        self.table_builder.create_summary_sheet("Summary", summary_data)

        # Evidence Matrix sheet
        headers = [
            "Control ID",
            "Criteria",
            "Title",
            "Evidence Description",
            "Evidence Type",
            "Collected Date",
            "Design Effectiveness",
            "Operating Effectiveness",
            "Exceptions",
            "Status",
        ]

        rows = []
        for record in evidence_records:
            rows.append(
                [
                    record.get("control_id", "N/A"),
                    self._format_status(record.get("criteria", "")),
                    record.get("title", "N/A"),
                    record.get("description", "N/A"),
                    record.get("evidence_type", "N/A"),
                    self._format_date(record.get("collected_at")),
                    self._format_status(record.get("design_effectiveness", "")),
                    self._format_status(record.get("operating_effectiveness", "")),
                    record.get("exceptions_noted", 0),
                    self._format_status(record.get("status", "")),
                ]
            )

        self.table_builder.create_evidence_sheet(
            sheet_name="Evidence Matrix",
            headers=headers,
            rows=rows,
            col_widths=[12, 18, 25, 40, 15, 15, 20, 20, 12, 15],
            status_column=9,  # Status column (0-indexed)
        )

        # Control Mappings sheet
        mappings = report_data.get("control_mappings", [])
        if mappings:
            mapping_headers = [
                "SOC 2 Control ID",
                "Guardrail Control ID",
                "Guardrail Control Name",
                "Mapping Rationale",
                "Coverage %",
                "Gaps",
            ]
            mapping_rows = []
            for m in mappings:
                mapping_data = m if isinstance(m, dict) else m.model_dump()
                mapping_rows.append(
                    [
                        mapping_data.get("soc2_control_id", "N/A"),
                        mapping_data.get("guardrail_control_id", "N/A"),
                        mapping_data.get("guardrail_control_name", "N/A"),
                        mapping_data.get("mapping_rationale", "N/A"),
                        mapping_data.get("coverage_percentage", 100),
                        ", ".join(mapping_data.get("gaps", [])) or "None",
                    ]
                )

            self.table_builder.create_evidence_sheet(
                sheet_name="Control Mappings",
                headers=mapping_headers,
                rows=mapping_rows,
                col_widths=[18, 20, 25, 45, 12, 30],
            )

        # Trust Service Criteria breakdown sheets
        criteria_sections = report_data.get("criteria_sections", [])
        for section in criteria_sections:
            criteria = section.get("criteria", "unknown")
            criteria_name = str(criteria).replace("_", " ").title()

            controls = section.get("controls", [])
            if controls:
                tsc_headers = ["Control ID", "Title", "Objective", "Guidance"]
                tsc_rows = []
                for ctrl in controls:
                    tsc_rows.append(
                        [
                            ctrl.get("control_id", "N/A"),
                            ctrl.get("title", "N/A"),
                            ctrl.get("control_objective", "N/A"),
                            ctrl.get("implementation_guidance", "N/A") or "N/A",
                        ]
                    )

                # Truncate sheet name to Excel's 31 char limit
                sheet_name = f"TSC - {criteria_name}"[:31]
                self.table_builder.create_evidence_sheet(
                    sheet_name=sheet_name,
                    headers=tsc_headers,
                    rows=tsc_rows,
                    col_widths=[15, 30, 40, 40],
                )

        # Generate output path if not provided
        if output_path is None:
            output_dir = _ensure_output_dir()
            timestamp = self._format_date(datetime.now(timezone.utc)).strftime("%Y%m%d_%H%M%S")
            output_path = output_dir / f"soc2_matrix_{timestamp}.xlsx"

        self._save_workbook(output_path)

        logger.info(f"Generated SOC 2 XLSX matrix: {output_path}")
        return Path(output_path)

    def generate_iso27001_matrix(
        self,
        report_data: DocumentData,
        output_path: Optional[Union[str, Path]] = None,
    ) -> Path:
        """
        Generate an ISO 27001 evidence matrix in XLSX format.

        Args:
            report_data: ISO 27001 report data dictionary.
            output_path: Optional output file path.

        Returns:
            Path to the generated XLSX file.
        """
        evidence_records = report_data.get("evidence_records", [])
        self._create_workbook(len(evidence_records))

        org_name = report_data.get("organization_name", "Organization")
        scope = report_data.get("isms_scope", "")

        # Summary sheet
        summary_data = [
            ("Report Type", "ISO 27001:2022 Annex A Evidence Matrix"),
            ("Organization", org_name),
            ("ISMS Scope", scope or "Not specified"),
            ("Generated At", self._format_date(datetime.now(timezone.utc))),
            ("Total Controls", report_data.get("total_controls", 93)),
            ("Applicable Controls", report_data.get("applicable_controls", 0)),
            ("Implemented Controls", report_data.get("implemented_controls", 0)),
            (
                "Implementation Percentage",
                f"{report_data.get('implementation_percentage', 0):.1f}%",
            ),
        ]
        self.table_builder.create_summary_sheet("Summary", summary_data)

        # Statement of Applicability sheet
        soa = report_data.get("statement_of_applicability", {})
        entries = soa.get("entries", [])
        if entries:
            soa_headers = [
                "Control ID",
                "Control Title",
                "Theme",
                "Applicable",
                "Justification",
                "Implementation Status",
                "Implementation Method",
                "Evidence Reference",
            ]
            soa_rows = []
            for entry in entries:
                soa_rows.append(
                    [
                        entry.get("control_id", "N/A"),
                        entry.get("control_title", "N/A"),
                        self._format_status(entry.get("theme", "")),
                        "Yes" if entry.get("applicability") in ["applicable", True] else "No",
                        entry.get("justification", "N/A") or "N/A",
                        self._format_status(entry.get("implementation_status", "")),
                        entry.get("implementation_method", "N/A") or "N/A",
                        entry.get("evidence_reference", "N/A") or "N/A",
                    ]
                )

            self.table_builder.create_evidence_sheet(
                sheet_name="Statement of Applicability",
                headers=soa_headers,
                rows=soa_rows,
                col_widths=[12, 35, 20, 12, 35, 20, 25, 25],
                status_column=5,  # Implementation Status column
            )

        # Evidence Matrix sheet
        if evidence_records:
            evidence_headers = [
                "Control ID",
                "Theme",
                "Evidence Description",
                "Evidence Type",
                "Source",
                "Collected Date",
                "Status",
                "Notes",
            ]
            evidence_rows = []
            for record in evidence_records:
                evidence_rows.append(
                    [
                        record.get("control_id", "N/A"),
                        self._format_status(record.get("theme", "")),
                        record.get("description", "N/A"),
                        record.get("evidence_type", "N/A"),
                        record.get("source", "N/A") or "N/A",
                        self._format_date(record.get("collected_at")),
                        self._format_status(record.get("status", "")),
                        record.get("notes", "N/A") or "N/A",
                    ]
                )

            self.table_builder.create_evidence_sheet(
                sheet_name="Evidence Matrix",
                headers=evidence_headers,
                rows=evidence_rows,
                col_widths=[12, 20, 40, 15, 20, 15, 15, 30],
                status_column=6,
            )

        # Theme breakdown sheets
        theme_sections = report_data.get("theme_sections", [])
        for section in theme_sections:
            theme = section.get("theme", "unknown")
            theme_name = str(theme).replace("_", " ").title()

            controls = section.get("controls", [])
            if controls:
                theme_headers = [
                    "Control ID",
                    "Title",
                    "Description",
                    "Status",
                ]
                theme_rows = []
                for ctrl in controls:
                    theme_rows.append(
                        [
                            ctrl.get("control_id", "N/A"),
                            ctrl.get("title", "N/A"),
                            ctrl.get("description", "N/A"),
                            self._format_status(ctrl.get("status", "")),
                        ]
                    )

                # Truncate sheet name to Excel's 31 char limit
                sheet_name = theme_name[:31]
                self.table_builder.create_evidence_sheet(
                    sheet_name=sheet_name,
                    headers=theme_headers,
                    rows=theme_rows,
                    col_widths=[12, 30, 50, 18],
                    status_column=3,
                )

        # Generate output path if not provided
        if output_path is None:
            output_dir = _ensure_output_dir()
            timestamp = self._format_date(datetime.now(timezone.utc)).strftime("%Y%m%d_%H%M%S")
            output_path = output_dir / f"iso27001_matrix_{timestamp}.xlsx"

        self._save_workbook(output_path)

        logger.info(f"Generated ISO 27001 XLSX matrix: {output_path}")
        return Path(output_path)

    def generate_gdpr_matrix(
        self,
        report_data: DocumentData,
        output_path: Optional[Union[str, Path]] = None,
    ) -> Path:
        """
        Generate a GDPR Article 30 records matrix in XLSX format.

        Args:
            report_data: GDPR report data dictionary.
            output_path: Optional output file path.

        Returns:
            Path to the generated XLSX file.
        """
        processing_activities = []
        controller_record = report_data.get("controller_record", {})
        if controller_record:
            processing_activities = controller_record.get("processing_activities", [])

        self._create_workbook(len(processing_activities))

        org_name = report_data.get("organization_name", "Organization")
        entity_role = report_data.get("entity_role", "controller")

        # Summary sheet
        summary_data = [
            ("Report Type", "GDPR Article 30 Records of Processing"),
            ("Organization", org_name),
            ("Entity Role", self._format_status(entity_role)),
            ("Generated At", self._format_date(datetime.now(timezone.utc))),
            ("Total Processing Activities", len(processing_activities)),
            ("Controller", controller_record.get("controller_name", "N/A")),
        ]

        # Add DPO info if available
        dpo = controller_record.get("dpo", {})
        if dpo:
            summary_data.append(("DPO Name", dpo.get("name", "N/A")))
            summary_data.append(("DPO Email", dpo.get("email", "N/A")))

        self.table_builder.create_summary_sheet("Summary", summary_data)

        # Processing Activities sheet (Article 30(1))
        if processing_activities:
            pa_headers = [
                "Activity Name",
                "Processing ID",
                "Purposes",
                "Lawful Basis",
                "Data Subject Categories",
                "Personal Data Categories",
                "Recipients",
                "Third Country Transfers",
                "Retention Period",
                "Security Measures",
                "Status",
            ]
            pa_rows = []
            for activity in processing_activities:
                purposes = activity.get("purposes", [])
                purposes_str = ", ".join(purposes[:3])
                if len(purposes) > 3:
                    # PERFORMANCE: Consider using list.append() + "".join() instead of += in loops
                    purposes_str += f" (+{len(purposes) - 3} more)"

                recipients = activity.get("recipients", [])
                recipients_str = ", ".join(
                    [
                        r.get("name", str(r)) if isinstance(r, dict) else str(r)
                        for r in recipients[:2]
                    ]
                )
                if len(recipients) > 2:
                    # PERFORMANCE: Consider using list.append() + "".join() instead of += in loops
                    recipients_str += f" (+{len(recipients) - 2} more)"

                data_subjects = activity.get("data_subject_categories", [])
                data_categories = activity.get("personal_data_categories", [])

                pa_rows.append(
                    [
                        activity.get("name", "N/A"),
                        activity.get("processing_id", "N/A"),
                        purposes_str or "N/A",
                        self._format_status(activity.get("lawful_basis", "")),
                        ", ".join(data_subjects[:2]) if data_subjects else "N/A",
                        ", ".join(data_categories[:2]) if data_categories else "N/A",
                        recipients_str or "N/A",
                        "Yes" if activity.get("third_country_transfers") else "No",
                        activity.get("retention_period", "N/A") or "N/A",
                        activity.get("security_measures_summary", "N/A") or "N/A",
                        self._format_status(activity.get("status", "")),
                    ]
                )

            self.table_builder.create_evidence_sheet(
                sheet_name="Processing Activities",
                headers=pa_headers,
                rows=pa_rows,
                col_widths=[25, 15, 30, 18, 25, 25, 25, 15, 20, 25, 15],
                status_column=10,
            )

        # Data Flows sheet
        data_flows = report_data.get("data_flows", [])
        if data_flows:
            df_headers = [
                "Flow Name",
                "Flow ID",
                "Data Source",
                "Data Destination",
                "Data Categories",
                "Crosses Border",
                "Transfer Mechanism",
                "Encryption in Transit",
            ]
            df_rows = []
            for flow in data_flows:
                categories = flow.get("data_categories", [])
                df_rows.append(
                    [
                        flow.get("name", "N/A"),
                        flow.get("flow_id", "N/A"),
                        flow.get("data_source", "N/A"),
                        flow.get("data_destination", "N/A"),
                        ", ".join(categories[:3]) if categories else "N/A",
                        "Yes" if flow.get("crosses_border") else "No",
                        flow.get("transfer_mechanism", "N/A") or "N/A",
                        "Yes" if flow.get("encrypted_in_transit") else "No",
                    ]
                )

            self.table_builder.create_evidence_sheet(
                sheet_name="Data Flows",
                headers=df_headers,
                rows=df_rows,
                col_widths=[25, 15, 25, 25, 30, 15, 25, 18],
            )

        # Security Measures sheet
        security_measures = report_data.get("security_measures", [])
        if security_measures:
            sm_headers = [
                "Measure ID",
                "Category",
                "Description",
                "Implementation Status",
                "Last Reviewed",
            ]
            sm_rows = []
            for measure in security_measures:
                measure_data = (
                    measure if isinstance(measure, dict) else {"description": str(measure)}
                )
                sm_rows.append(
                    [
                        measure_data.get("measure_id", "N/A"),
                        self._format_status(measure_data.get("category", "")),
                        measure_data.get("description", str(measure)),
                        self._format_status(measure_data.get("status", "implemented")),
                        self._format_date(measure_data.get("last_reviewed")),
                    ]
                )

            self.table_builder.create_evidence_sheet(
                sheet_name="Security Measures",
                headers=sm_headers,
                rows=sm_rows,
                col_widths=[15, 20, 50, 20, 15],
                status_column=3,
            )

        # Generate output path if not provided
        if output_path is None:
            output_dir = _ensure_output_dir()
            timestamp = self._format_date(datetime.now(timezone.utc)).strftime("%Y%m%d_%H%M%S")
            output_path = output_dir / f"gdpr_matrix_{timestamp}.xlsx"

        self._save_workbook(output_path)

        logger.info(f"Generated GDPR XLSX matrix: {output_path}")
        return Path(output_path)

    def generate_euaiact_matrix(
        self,
        report_data: DocumentData,
        output_path: Optional[Union[str, Path]] = None,
    ) -> Path:
        """
        Generate an EU AI Act compliance matrix in XLSX format.

        Args:
            report_data: EU AI Act report data dictionary.
            output_path: Optional output file path.

        Returns:
            Path to the generated XLSX file.
        """
        ai_systems = report_data.get("ai_systems", [])
        self._create_workbook(len(ai_systems))

        org_name = report_data.get("organization_name", "Organization")
        org_role = report_data.get("organization_role", "provider")

        # Summary sheet
        high_risk = report_data.get("high_risk_systems_count", 0)
        limited_risk = report_data.get("limited_risk_systems_count", 0)
        minimal_risk = report_data.get("minimal_risk_systems_count", 0)

        summary_data = [
            ("Report Type", "EU AI Act Compliance Matrix"),
            ("Organization", org_name),
            ("Organization Role", self._format_status(org_role)),
            ("Generated At", self._format_date(datetime.now(timezone.utc))),
            ("Total AI Systems", len(ai_systems)),
            ("High-Risk Systems", high_risk),
            ("Limited-Risk Systems", limited_risk),
            ("Minimal-Risk Systems", minimal_risk),
        ]
        self.table_builder.create_summary_sheet("Summary", summary_data)

        # AI Systems Inventory sheet
        if ai_systems:
            sys_headers = [
                "System ID",
                "System Name",
                "Risk Level",
                "High-Risk Category",
                "Intended Purpose",
                "Provider/Deployer",
                "Compliance Status",
                "Last Assessment",
            ]
            sys_rows = []
            for system in ai_systems:
                sys_rows.append(
                    [
                        system.get("system_id", "N/A"),
                        system.get("system_name", "N/A"),
                        self._format_status(system.get("risk_level", "")),
                        system.get("high_risk_category", "N/A") or "N/A",
                        (
                            system.get("intended_purpose", "N/A")[:50] + "..."
                            if len(system.get("intended_purpose", "")) > 50
                            else system.get("intended_purpose", "N/A")
                        ),
                        self._format_status(system.get("provider_role", "")),
                        self._format_status(system.get("compliance_status", "")),
                        self._format_date(system.get("last_assessment_date")),
                    ]
                )

            self.table_builder.create_evidence_sheet(
                sheet_name="AI Systems Inventory",
                headers=sys_headers,
                rows=sys_rows,
                col_widths=[15, 25, 15, 20, 35, 18, 18, 15],
                status_column=6,
            )

        # Risk Assessments sheet
        risk_assessments = report_data.get("risk_assessments", [])
        if risk_assessments:
            ra_headers = [
                "Assessment ID",
                "System Name",
                "Risk Level",
                "High-Risk Category",
                "Assessment Date",
                "Assessor",
                "Key Findings",
                "Recommendations",
            ]
            ra_rows = []
            for assessment in risk_assessments:
                ra_rows.append(
                    [
                        assessment.get("assessment_id", "N/A"),
                        assessment.get("system_name", "N/A"),
                        self._format_status(assessment.get("risk_level", "")),
                        assessment.get("high_risk_category", "N/A") or "N/A",
                        self._format_date(assessment.get("assessment_date")),
                        assessment.get("assessor", "N/A") or "N/A",
                        assessment.get("key_findings", "N/A") or "N/A",
                        assessment.get("recommendations", "N/A") or "N/A",
                    ]
                )

            self.table_builder.create_evidence_sheet(
                sheet_name="Risk Assessments",
                headers=ra_headers,
                rows=ra_rows,
                col_widths=[15, 25, 15, 20, 15, 20, 35, 35],
            )

        # Conformity Assessments sheet
        conformity_assessments = report_data.get("conformity_assessments", [])
        if conformity_assessments:
            ca_headers = [
                "System ID",
                "Assessment Type",
                "Assessment Date",
                "Result",
                "Certificate Number",
                "Expiry Date",
                "Notified Body",
                "Notes",
            ]
            ca_rows = []
            for ca in conformity_assessments:
                ca_rows.append(
                    [
                        ca.get("system_id", "N/A"),
                        self._format_status(ca.get("assessment_type", "")),
                        self._format_date(ca.get("assessment_date")),
                        self._format_status(ca.get("assessment_result", "")),
                        ca.get("certificate_number", "N/A") or "N/A",
                        self._format_date(ca.get("expiry_date")),
                        ca.get("notified_body", "N/A") or "N/A",
                        ca.get("notes", "N/A") or "N/A",
                    ]
                )

            self.table_builder.create_evidence_sheet(
                sheet_name="Conformity Assessments",
                headers=ca_headers,
                rows=ca_rows,
                col_widths=[15, 22, 15, 15, 20, 15, 25, 30],
                status_column=3,
            )

        # Technical Documentation Checklist (Annex IV)
        tech_docs = report_data.get("technical_documentation", [])
        if tech_docs:
            td_headers = [
                "System Name",
                "General Description",
                "Elements & Development",
                "Monitoring & Functioning",
                "Risk Management",
                "Data Governance",
                "Human Oversight",
                "Overall Status",
            ]
            td_rows = []
            for doc in tech_docs:
                reqs = doc.get("annex_iv_requirements", {})
                td_rows.append(
                    [
                        doc.get("system_name", "N/A"),
                        "Complete" if reqs.get("general_description") else "Incomplete",
                        "Complete" if reqs.get("elements_development") else "Incomplete",
                        "Complete" if reqs.get("monitoring_functioning") else "Incomplete",
                        "Complete" if reqs.get("risk_management") else "Incomplete",
                        "Complete" if reqs.get("data_governance") else "Incomplete",
                        "Complete" if reqs.get("human_oversight") else "Incomplete",
                        self._format_status(doc.get("documentation_status", "")),
                    ]
                )

            self.table_builder.create_evidence_sheet(
                sheet_name="Technical Documentation",
                headers=td_headers,
                rows=td_rows,
                col_widths=[25, 18, 20, 20, 18, 18, 18, 15],
                status_column=7,
            )

        # Generate output path if not provided
        if output_path is None:
            output_dir = _ensure_output_dir()
            timestamp = self._format_date(datetime.now(timezone.utc)).strftime("%Y%m%d_%H%M%S")
            output_path = output_dir / f"euaiact_matrix_{timestamp}.xlsx"

        self._save_workbook(output_path)

        logger.info(f"Generated EU AI Act XLSX matrix: {output_path}")
        return Path(output_path)


def generate_xlsx(
    report_data: DocumentData,
    framework: Union[str, ComplianceFramework],
    output_path: Optional[Union[str, Path]] = None,
    write_only: bool = False,
) -> Path:
    """
    Generate an XLSX compliance evidence matrix for the specified framework.

    This is the main entry point for XLSX generation.

    Args:
        report_data: Report data dictionary containing all compliance information.
        framework: Compliance framework (soc2, iso27001, gdpr, euaiact).
        output_path: Optional output file path. If not provided, a default path is used.
        write_only: Enable write_only mode for large files. Automatically enabled
                   when data exceeds 10,000 rows.

    Returns:
        Path to the generated XLSX file.

    Raises:
        ValueError: If the framework is not supported.

    Example:
        >>> from acgs2.services.compliance_docs.src.generators.xlsx_generator import (
        ...     generate_xlsx
        ... )
        >>> path = generate_xlsx(
        ...     report_data={"organization_name": "Acme Corp", ...},
        ...     framework="soc2",
        ... )
        >>> f"Matrix generated at: {path}")
    """
    generator = XLSXGenerator(write_only=write_only)

    # Normalize framework
    if isinstance(framework, ComplianceFramework):
        framework_str = framework.value
    else:
        framework_str = str(framework).lower()

    if framework_str == "soc2":
        return generator.generate_soc2_matrix(report_data, output_path)
    elif framework_str == "iso27001":
        return generator.generate_iso27001_matrix(report_data, output_path)
    elif framework_str == "gdpr":
        return generator.generate_gdpr_matrix(report_data, output_path)
    elif framework_str == "euaiact":
        return generator.generate_euaiact_matrix(report_data, output_path)
    else:
        raise ValueError(
            f"Unsupported framework: {framework}. "
            f"Supported frameworks: soc2, iso27001, gdpr, euaiact"
        )


def generate_xlsx_to_buffer(
    report_data: DocumentData,
    framework: Union[str, ComplianceFramework],
    write_only: bool = False,
) -> io.BytesIO:
    """
    Generate an XLSX compliance evidence matrix to an in-memory buffer.

    This is useful for streaming responses without writing to disk.

    Args:
        report_data: Report data dictionary containing all compliance information.
        framework: Compliance framework (soc2, iso27001, gdpr, euaiact).
        write_only: Enable write_only mode for large files.

    Returns:
        BytesIO buffer containing the generated XLSX.

    Raises:
        ValueError: If the framework is not supported.

    Example:
        >>> buffer = generate_xlsx_to_buffer(report_data, "soc2")
        >>> # Use buffer.getvalue() to get bytes for streaming
    """
    buffer = io.BytesIO()
    generator = XLSXGenerator(write_only=write_only)

    # Normalize framework
    if isinstance(framework, ComplianceFramework):
        framework_str = framework.value
    else:
        framework_str = str(framework).lower()

    # Create workbook and generate content
    org_name = report_data.get("organization_name", "Organization")

    if framework_str == "soc2":
        generator._create_workbook()
        summary_data = [
            ("Report Type", "SOC 2 Type II Evidence Matrix"),
            ("Organization", org_name),
            ("Generated At", generator._format_date(datetime.now(timezone.utc))),
        ]
        generator.table_builder.create_summary_sheet("Summary", summary_data)
        generator.table_builder.create_evidence_sheet(
            sheet_name="Evidence Matrix",
            headers=["Control ID", "Criteria", "Description", "Status"],
            rows=[["See generate_xlsx() for full implementation", "", "", ""]],
        )
    elif framework_str == "iso27001":
        generator._create_workbook()
        summary_data = [
            ("Report Type", "ISO 27001:2022 Annex A Evidence Matrix"),
            ("Organization", org_name),
            ("Generated At", generator._format_date(datetime.now(timezone.utc))),
        ]
        generator.table_builder.create_summary_sheet("Summary", summary_data)
        generator.table_builder.create_evidence_sheet(
            sheet_name="Evidence Matrix",
            headers=["Control ID", "Theme", "Description", "Status"],
            rows=[["See generate_xlsx() for full implementation", "", "", ""]],
        )
    elif framework_str == "gdpr":
        generator._create_workbook()
        summary_data = [
            ("Report Type", "GDPR Article 30 Records"),
            ("Organization", org_name),
            ("Generated At", generator._format_date(datetime.now(timezone.utc))),
        ]
        generator.table_builder.create_summary_sheet("Summary", summary_data)
        generator.table_builder.create_evidence_sheet(
            sheet_name="Processing Activities",
            headers=["Activity", "Purposes", "Data Categories", "Status"],
            rows=[["See generate_xlsx() for full implementation", "", "", ""]],
        )
    elif framework_str == "euaiact":
        generator._create_workbook()
        summary_data = [
            ("Report Type", "EU AI Act Compliance Matrix"),
            ("Organization", org_name),
            ("Generated At", generator._format_date(datetime.now(timezone.utc))),
        ]
        generator.table_builder.create_summary_sheet("Summary", summary_data)
        generator.table_builder.create_evidence_sheet(
            sheet_name="AI Systems",
            headers=["System ID", "Name", "Risk Level", "Status"],
            rows=[["See generate_xlsx() for full implementation", "", "", ""]],
        )
    else:
        raise ValueError(
            f"Unsupported framework: {framework}. "
            f"Supported frameworks: soc2, iso27001, gdpr, euaiact"
        )

    generator._save_workbook(buffer)
    buffer.seek(0)
    return buffer
